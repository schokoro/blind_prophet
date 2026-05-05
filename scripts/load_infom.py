import argparse
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from amnesiac.store import apply_migrations, get_connection


SOURCE_URL = "https://www.cbr.ru/Collection/Collection/File/60881/Infl_exp_26-04.xlsx"
SHEET_NAME = "Данные за все годы"
EXPECTED_LABEL = "ожидаемая инфляция (в %)"
DEFAULT_DB_PATH = Path("data/db/blind_prophet.db")


def download_xlsx(url: str) -> Path:
    try:
        import requests
    except ModuleNotFoundError as exc:
        raise RuntimeError("Missing dependency: requests. Install requirements.txt.") from exc

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        try:
            response = requests.get(url, timeout=60)
            response.raise_for_status()
            tmp.write(response.content)
        except requests.RequestException as exc:
            tmp_path.unlink(missing_ok=True)
            raise RuntimeError(f"Download failed: {exc}") from exc
        except OSError:
            tmp_path.unlink(missing_ok=True)
            raise
        return tmp_path


def extract_rows(path: Path) -> list[tuple[str, float]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install requirements.txt.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[SHEET_NAME]
        dates = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        values = next(sheet.iter_rows(min_row=71, max_row=71, values_only=True))

        label = values[0]
        if label != EXPECTED_LABEL:
            raise ValueError(
                f"Unexpected row 71 label: {label!r}; expected {EXPECTED_LABEL!r}"
            )

        rows = []
        for survey_date, median_12m in zip(dates[1:], values[1:]):
            if isinstance(survey_date, datetime) and isinstance(median_12m, (int, float)):
                rows.append((survey_date.date().isoformat(), round(float(median_12m), 4)))
        return rows
    finally:
        workbook.close()


def load_rows(db_path: Path, rows: list[tuple[str, float]]) -> tuple[int, int]:
    conn = get_connection(db_path)
    try:
        apply_migrations(conn)
        loaded = 0
        for row in rows:
            cursor = conn.execute(
                """
                INSERT OR IGNORE INTO infom_expectations (
                    survey_date, median_12m, source_url
                )
                VALUES (?, ?, ?)
                """,
                (row[0], row[1], SOURCE_URL),
            )
            loaded += cursor.rowcount
        conn.commit()
        return loaded, len(rows) - loaded
    finally:
        conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load CBR InFOM inflation expectations into SQLite."
    )
    parser.add_argument(
        "--db-path",
        type=Path,
        default=Path(os.environ.get("DB_PATH", DEFAULT_DB_PATH)),
        help="SQLite DB path. Defaults to DB_PATH or data/db/blind_prophet.db.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    tmp_path: Path | None = None

    try:
        tmp_path = download_xlsx(SOURCE_URL)
        rows = extract_rows(tmp_path)
        loaded, skipped = load_rows(args.db_path, rows)
    except (RuntimeError, FileNotFoundError, OSError, ValueError, KeyError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    finally:
        if tmp_path is not None:
            tmp_path.unlink(missing_ok=True)

    print(f"Loaded {loaded} rows, skipped {skipped} duplicates")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
